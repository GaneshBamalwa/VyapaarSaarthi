from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import io
import logging

from app.database.session import get_db
from app.models.expense import ExpenseEntry, MonthlyReportMeta
from app.models.order import Order
from app.schemas.expense import (
    ExpenseCreateSchema, ExpenseResponseSchema, MonthlyReportSchema, CategoryBreakdownSchema
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api", tags=["Expenses"])
logger = logging.getLogger(__name__)

def get_current_month():
    return datetime.utcnow().strftime("%Y-%m")

@router.get("/expenses", response_model=List[ExpenseResponseSchema])
async def list_expenses(month: Optional[str] = None, db: Session = Depends(get_db)):
    target_month = month or get_current_month()
    expenses = db.query(ExpenseEntry).filter(
        ExpenseEntry.month == target_month,
        ExpenseEntry.is_deleted == False
    ).all()
    return expenses

@router.post("/expenses", response_model=ExpenseResponseSchema)
async def create_expense(expense: ExpenseCreateSchema, db: Session = Depends(get_db)):
    target_month = get_current_month()
    
    meta = db.query(MonthlyReportMeta).filter_by(month=target_month).first()
    if meta and meta.is_closed:
        raise HTTPException(status_code=400, detail={"error": "Month is closed", "code": 400})
        
    db_expense = ExpenseEntry(
        month=target_month,
        category=expense.category.value,
        description=expense.description,
        vendor_name=expense.vendor_name,
        amount=expense.amount,
        due_date=expense.due_date,
        linked_order_id=expense.linked_order_id,
        receipt_ref=expense.receipt_ref
    )
    db.add(db_expense)
    db.commit()
    db.refresh(db_expense)
    logger.info(f"[EXPENSE] Created new expense {db_expense.id} for {target_month}")
    return db_expense

@router.patch("/expenses/{expense_id}/mark-paid")
async def mark_paid(expense_id: int, db: Session = Depends(get_db)):
    expense = db.query(ExpenseEntry).filter(ExpenseEntry.id == expense_id, ExpenseEntry.is_deleted == False).first()
    if not expense:
        raise HTTPException(status_code=404, detail={"error": "Expense not found", "code": 404})
        
    meta = db.query(MonthlyReportMeta).filter_by(month=expense.month).first()
    if meta and meta.is_closed:
        raise HTTPException(status_code=400, detail={"error": "Month is closed", "code": 400})
        
    expense.is_paid = True
    expense.paid_on = datetime.utcnow().strftime("%Y-%m-%d")
    db.commit()
    db.refresh(expense)
    logger.info(f"[EXPENSE] Marked expense {expense.id} as paid")
    return {"status": "success", "paid_on": expense.paid_on}

@router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: int, db: Session = Depends(get_db)):
    expense = db.query(ExpenseEntry).filter(ExpenseEntry.id == expense_id, ExpenseEntry.is_deleted == False).first()
    if not expense:
        raise HTTPException(status_code=404, detail={"error": "Expense not found", "code": 404})
        
    meta = db.query(MonthlyReportMeta).filter_by(month=expense.month).first()
    if meta and meta.is_closed:
        raise HTTPException(status_code=400, detail={"error": "Month is closed", "code": 400})
        
    expense.is_deleted = True
    db.commit()
    logger.info(f"[EXPENSE] Deleted expense {expense.id}")
    return {"status": "success"}

@router.get("/reports/monthly", response_model=MonthlyReportSchema)
async def get_monthly_report(month: Optional[str] = None, db: Session = Depends(get_db)):
    target_month = month or get_current_month()
    
    meta = db.query(MonthlyReportMeta).filter_by(month=target_month).first()
    
    expenses = db.query(ExpenseEntry).filter(
        ExpenseEntry.month == target_month,
        ExpenseEntry.is_deleted == False
    ).all()
    
    expenses_total = sum(e.amount for e in expenses)
    
    orders = db.query(Order).filter(Order.delivery_date.between(f"{target_month}-01", f"{target_month}-31")).all()
    orders_total = 0
    for order in orders:
        if order.items:
            orders_total += sum(i.price * i.quantity for i in order.items if i.price and i.quantity)
            
    net_position = orders_total - expenses_total
    
    categories = {}
    for e in expenses:
        categories[e.category] = categories.get(e.category, 0) + e.amount
        
    breakdown = [CategoryBreakdownSchema(category=k, amount=v) for k, v in categories.items()]
    
    return MonthlyReportSchema(
        month=target_month,
        orders_total=orders_total,
        expenses_total=expenses_total,
        net_position=net_position,
        is_closed=meta.is_closed if meta else False,
        category_breakdown=breakdown
    )

def _style_header(ws):
    header_font = Font(name='Arial', bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def _auto_fit_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 4, 40)
        ws.column_dimensions[column].width = adjusted_width

@router.get("/reports/download")
async def download_report(month: Optional[str] = None, report_type: Optional[str] = "expenses", db: Session = Depends(get_db)):
    target_month = month or get_current_month()
    year, month_str = target_month.split("-")
    
    buffer = generate_report_excel(db, target_month, report_type)
    
    filename = f"{report_type}_{year}_{month_str}_vyapaarsaarthi.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

def generate_report_excel(db: Session, target_month: str, report_type: str) -> io.BytesIO:
    year, month_str = target_month.split("-")
    
    orders = db.query(Order).filter(Order.delivery_date.between(f"{target_month}-01", f"{target_month}-31")).all()
    expenses = db.query(ExpenseEntry).filter(ExpenseEntry.month == target_month, ExpenseEntry.is_deleted == False).all()
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: Expenses (Default)
    ws_exp = wb.active
    ws_exp.title = "Expenses"
    ws_exp.append(["Expense ID", "Category", "Description", "Vendor", "Amount (\u20B9)", "Is Paid", "Due Date", "Paid On", "Linked Order ID", "Receipt Ref"])
    
    alt_fill = PatternFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    paid_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    due_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    overdue_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    exp_row_idx = 2
    total_expenses = 0.0
    for idx, exp in enumerate(expenses):
        is_paid_str = "✓ PAID" if exp.is_paid else "✗ PENDING"
        ws_exp.append([
            exp.id, exp.category, exp.description, exp.vendor_name, exp.amount, is_paid_str,
            exp.due_date, exp.paid_on, exp.linked_order_id, exp.receipt_ref
        ])
        
        for cell in ws_exp[exp_row_idx]:
            cell.font = Font(name='Arial')
            cell.fill = alt_fill if idx % 2 != 0 else white_fill
            
        ws_exp[f"E{exp_row_idx}"].number_format = '\u20B9#,##0.00'
        ws_exp[f"F{exp_row_idx}"].fill = paid_fill if exp.is_paid else overdue_fill
        total_expenses += float(exp.amount or 0)
        
        exp_row_idx += 1
        
    ws_exp.append(["TOTAL EXPENSES", "", "", "", total_expenses, "", "", "", "", ""])
    for cell in ws_exp[exp_row_idx]:
        cell.font = Font(name='Arial', bold=True)
        cell.fill = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
    
    # Sheet 2: Orders
    ws_orders = wb.create_sheet(title="Orders")
    ws_orders.append(["Order ID", "Customer Name", "Product", "Quantity", "Unit Price (\u20B9)", "Total Value (\u20B9)", "Delivery Date", "Payment Status"])
    
    order_row_idx = 2
    total_order_qty = 0
    total_order_value = 0.0
    for idx, order in enumerate(orders):
        product_name = order.items[0].name if order.items else ""
        qty = order.items[0].quantity if order.items else 0
        price = order.items[0].price if order.items else 0
        total = qty * price
        
        status = "OVERDUE"
        status_fill = overdue_fill
        if order.status.value == "COMPLETED":
            status = "PAID"
            status_fill = paid_fill
        elif order.delivery_date and order.delivery_date > datetime.utcnow().strftime("%Y-%m-%d"):
            status = "DUE ON DELIVERY"
            status_fill = due_fill
            
        ws_orders.append([
            order.id, order.customer, product_name, qty, price, total, order.delivery_date, status
        ])
        
        for cell in ws_orders[order_row_idx]:
            cell.font = Font(name='Arial')
            cell.fill = alt_fill if idx % 2 != 0 else white_fill
        
        ws_orders[f"E{order_row_idx}"].number_format = '\u20B9#,##0.00'
        ws_orders[f"F{order_row_idx}"].number_format = '\u20B9#,##0.00'
        ws_orders[f"H{order_row_idx}"].fill = status_fill
        
        total_order_qty += qty
        total_order_value += total
        
        order_row_idx += 1
        
    ws_orders.append(["MONTHLY TOTAL", "", "", total_order_qty, "", total_order_value, "", ""])
    for cell in ws_orders[order_row_idx]:
        cell.font = Font(name='Arial', bold=True)
        cell.fill = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
    
    _style_header(ws_orders)
    _auto_fit_columns(ws_orders)
    
    sum_start = exp_row_idx + 2
    accent_border = Border(left=Side(style='thick', color='1E3A5F'))
    
    ws_exp.cell(row=sum_start, column=1, value="Total Revenue (Orders):")
    ws_exp.cell(row=sum_start, column=2, value=total_order_value)
    ws_exp[f"A{sum_start}"].font = Font(name='Arial', bold=True)
    ws_exp[f"A{sum_start}"].border = accent_border
    ws_exp[f"B{sum_start}"].number_format = '\u20B9#,##0.00'
    
    ws_exp.cell(row=sum_start+1, column=1, value="Total Expenses:")
    ws_exp.cell(row=sum_start+1, column=2, value=total_expenses)
    ws_exp[f"A{sum_start+1}"].font = Font(name='Arial', bold=True)
    ws_exp[f"A{sum_start+1}"].border = accent_border
    ws_exp[f"B{sum_start+1}"].number_format = '\u20B9#,##0.00'
    
    ws_exp.cell(row=sum_start+2, column=1, value="Net Position:")
    ws_exp.cell(row=sum_start+2, column=2, value=total_order_value - total_expenses)
    ws_exp[f"A{sum_start+2}"].font = Font(name='Arial', bold=True)
    ws_exp[f"A{sum_start+2}"].border = accent_border
    ws_exp[f"B{sum_start+2}"].number_format = '\u20B9#,##0.00'
    
    _style_header(ws_exp)
    _auto_fit_columns(ws_exp)
    
    if report_type == "orders":
        wb.remove(ws_exp)
        wb.active = wb.sheetnames.index("Orders")
    elif report_type == "expenses":
        wb.remove(ws_orders)
        wb.active = wb.sheetnames.index("Expenses")
    else:
        # For "both" or any other fallback
        wb.active = wb.sheetnames.index("Expenses")
    
    import tempfile
    import subprocess
    import os
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_in:
        wb.save(tmp_in.name)
        tmp_in_path = tmp_in.name
        
    tmp_out_path = tmp_in_path.replace(".xlsx", "_recalc.xlsx")
    
    # Call recalc.py
    script_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "..", "scripts", "recalc.py")
    try:
        subprocess.run(["python", script_path, tmp_in_path, tmp_out_path], check=True)
        if os.path.exists(tmp_out_path):
            with open(tmp_out_path, "rb") as f:
                buffer = io.BytesIO(f.read())
        else:
            with open(tmp_in_path, "rb") as f:
                buffer = io.BytesIO(f.read())
    except Exception as e:
        logger.error(f"Failed to recalculate excel: {e}")
        with open(tmp_in_path, "rb") as f:
            buffer = io.BytesIO(f.read())
            
    # Cleanup temp files
    if os.path.exists(tmp_in_path):
        os.remove(tmp_in_path)
    if os.path.exists(tmp_out_path):
        os.remove(tmp_out_path)
    
    buffer.seek(0)
    return buffer
